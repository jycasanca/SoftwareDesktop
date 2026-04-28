class ExcelIO:
    def _estilo_cabecera(self, celda):
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            celda.fill = PatternFill("solid", fgColor="333333")
            celda.font = Font(color="FFFFFF", bold=True, name="Consolas", size=11)
            celda.alignment = Alignment(horizontal="center")
        except:
            pass
    
    def exportar_plan_contable(self, ruta, datos):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Plan Contable"
            
            cabeceras = ['codigo', 'descripcion', 'tipo_cuenta', 'naturaleza', 'activa']
            for i, cab in enumerate(cabeceras, 1):
                celda = ws.cell(row=1, column=i)
                celda.value = cab
                self._estilo_cabecera(celda)
            
            for fila in datos:
                ws.append([fila['codigo'], fila['descripcion'], fila['tipo_cuenta'],
                          fila['naturaleza'], fila['activa']])
            
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            wb.save(ruta)
            return True
        except Exception as e:
            print(f"Error exportando Excel: {e}")
            return False
    
    def importar_plan_contable(self, ruta):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta)
            ws = wb.active
            
            cabeceras_esperadas = ['codigo', 'descripcion', 'tipo_cuenta', 'naturaleza']
            tipos_validos = {'activo_corriente', 'activo_no_corriente', 'pasivo_corriente',
                           'pasivo_no_corriente', 'patrimonio', 'ingreso', 'gasto', 'costo'}
            naturalezas_validas = {'deudora', 'acreedora'}
            
            # Validar cabeceras
            fila1 = [cell.value for cell in ws[1]]
            for esperada in cabeceras_esperadas:
                if esperada not in fila1:
                    return [], [f"Falta columna: {esperada}"]
            
            registros = []
            errores = []
            
            for i, row in enumerate(ws.iter_rows(min_row=2), 2):
                try:
                    codigo = int(row[0].value) if row[0].value else None
                    descripcion = row[1].value
                    tipo_cuenta = row[2].value
                    naturaleza = row[3].value
                    
                    if not codigo:
                        errores.append(f"Fila {i}: Código vacío")
                        continue
                    
                    if tipo_cuenta not in tipos_validos:
                        errores.append(f"Fila {i}: Tipo cuenta inválido")
                        continue
                    
                    if naturaleza not in naturalezas_validas:
                        errores.append(f"Fila {i}: Naturaleza inválida")
                        continue
                    
                    registros.append({
                        'codigo': codigo,
                        'descripcion': descripcion,
                        'tipo_cuenta': tipo_cuenta,
                        'naturaleza': naturaleza,
                        'activa': 1
                    })
                except Exception as e:
                    errores.append(f"Fila {i}: {str(e)}")
            
            return registros, errores
        except Exception as e:
            return [], [f"Error leyendo archivo: {str(e)}"]
